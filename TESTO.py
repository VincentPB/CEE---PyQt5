import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import datetime

#=============================================================#

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

#==========================================================================#

doc='DONE/TR_TRA-EQ-111 2 SLXI 300.50.xlsx'


def switchOperation(filename): #Applique le traitement correspondant au fichier importé.

    header=str(pd.read_excel(filename, nrows=4))

    if(('115+103' in header) or ('103+115' in header)):
        deboublonner(filename, [0, 1], 'TRA EQ 115-103')
        print('TRAITEMENT TERMINE')        
    elif (("EQ-115" in header) or ("EQ-15" in header)):
        deboublonner(filename,[0, 1], 'TRA EQ 115')
        print('TRAITEMENT TERMINE')
    elif (("EQ-119" in header) or ("EQ-19" in header)):
        deboublonner(filename, [2, 1], 'TRA EQ 119')
        print('TRAITEMENT TERMINE')
    elif (("EQ-103" in header) or ("EQ-03" in header)):
        deboublonner(filename, INFO)
        print('TRAITEMENT TERMINE')
    elif (("EQ-101" in header) or ("EQ-01" in header)):
        deboublonner(filename, [1, 7], 'TRA EQ 101')
        print('TRAITEMENT TERMINE')   
    elif (("EQ-111" in header) or ("EQ-11" in header)):
        deboublonner(filename, [0, 6], 'TRA EQ 111')
        print('TRAITEMENT TERMINE')
    elif (("SE-113" in header) or ("SE-13" in header)):
        deboublonner(filename, [1, 3], 'TRA SE 113')
        print('TRAITEMENT TERMINE')
    elif (("SE-108" in header) or ("SE-08" in header)):
        deboublonner(filename, [1, 5], 'TRA SE 108')
        print('TRAITEMENT TERMINE')
    elif (("SE-105" in header) or ("SE-05" in header)):
        deboublonner(filename, [4, 3], 'TRA SE 105')
        print('TRAITEMENT TERMINE')
    elif (("SE-101" in header) or ("SE-01" in header)):
        deboublonner(filename, [1, 0, 3, 7], 'TRA SE 101')
        print('TRAITEMENT TERMINE')
    else:
        return("OPERATION INVALIDE")

def deboublonner(doc, indCrit, titre):

    Temps = datetime.datetime.now()
    TempsMax = datetime.datetime(Temps.year-10, Temps.month, Temps.day)
    header=pd.read_excel(doc, nrows=4)
    d=pd.read_excel(doc, header=5)
    NbRow = d.shape[0]
    NbCol = d.shape[1]
    ListCrit1=[]
    ListCritDate=[]
    ListeDoublons = []

    for i in range(NbRow):
        ListCrit1.append(d.iloc[i,indCrit[0]])
        ListCritDate.append(d.iloc[i,indCrit[1]])

    for i in range(NbRow-1):
        if(ListCrit1[i] in (ListCrit1[:i]+ListCrit1[i+1:]) or ListCritDate[i]<TempsMax):
            ListeDoublons.append(i)

    for ind in ListeDoublons:
        d=d.drop(ind)

    PostTra = xlsxwriter.Workbook('DEDOUBLONNE.xlsx')
    fueillasse = PostTra.add_worksheet(titre)

    for h1 in range(header.shape[0]):
        for h2 in range(header.shape[1]):
            if(str(header.iloc[h1,h2])!='nan'):
                fueillasse.write(h1, h2, header.iloc[h1,h2])

    PostTra.close()
    append_df_to_excel('DEDOUBLONNE.xlsx', d, sheet_name=titre, startrow=5, index=False)

switchOperation(doc)


    

